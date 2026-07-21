"""
Vuelca los datos de la sesión del 24/06/2026 (Colegio Gregorio Canella)
en las hojas "Registro" y "Comentarios" de hoja_rapida_evaluacion.xlsx.

IMPORTANTE: cierra el archivo en LibreOffice/Excel antes de ejecutar.

Uso:
    cd /home/menunezg/PyCharmMiscProject/TFM/olibot
    python evaluacion/volcar_sesion_20260624.py
"""

from openpyxl import load_workbook
from openpyxl.styles import Alignment

XLSX = "evaluacion/hoja_rapida_evaluacion.xlsx"
FECHA = "24/06/2026"
SESION = 1

# ── Hoja "Registro" ──────────────────────────────────────────────────────────
# Columnas (en el mismo orden que el script generador):
#  Fecha | Sesión nº | Apodo | Modalidad | Pareja con | Grupo | Edad |
#  Nivel (color) | ¿1ª vez? | Min. atención | Nº distracciones |
#  Activ. propuestas | Activ. superadas | Contenidos nuevos |
#  Interv. adulto (nº) | Motivo interv. | "Otra vez"/reintentos |
#  Satisfacción | ¿Repetiría? | Autonomía | ¿Usó voz? | Nota rápida

REGISTRO = [
    # ── Héctor (primer niño del día, 9:31-9:49) ──────────────────────────
    (
        FECHA, SESION, "Héctor", "Individual", "", "OLIBOT", "6", "Rojo",
        "Sí", 18, 3, 5, 4,
        "Letras básicas (nivel rojo); trazos de pintura",
        3, "Técnico",
        "Sí (botón repetir, autónomo)",
        "Contento", "Sí", "Media", "Mal",
        "Niño prueba / dificultades. No espera tutorial. Completa 4 ejercicios de "
        "pintura. Prefiere OLIBOT al papel ('más divertido'). No le gusta que repitan "
        "'Muy bien otra vez'. Reducir preguntas de estado emocional. Quiere dibujo libre.",
    ),
    # ── Cayetana (10:02-10:19) ────────────────────────────────────────────
    (
        FECHA, SESION, "Cayetana", "Individual", "", "OLIBOT", "5", "Rojo",
        "Sí", 17, 1, 5, 3,
        "a, as, l",
        2, "Pedagógico",
        "No",
        "Contento", "Sí", "Alta", "No usó",
        "Nivel automático → rojo. Trabaja a, as, l. Sin miedo; ríe. Paciente, espera "
        "los estados. Se cansa al final y pinta. Selecciona dibujo por preferencia, "
        "no por categoría.",
    ),
    # ── Ángela (10:35-10:53) ──────────────────────────────────────────────
    (
        FECHA, SESION, "Ángela", "Individual", "", "OLIBOT", "4", "Rojo",
        "Sí", 18, 1, 3, 1,
        "A (nivel rojo; ella dice ya sabérsela). Le gustaría matemáticas.",
        2, "Técnico",
        "No",
        "Contento", "Sí", "Alta", "No usó",
        "4 años, nivel rojo. Atiende bien el tutorial la 1ª vez; después lo ignora "
        "(pq se lo sabe). Va a pintar a los 5 min. Aprende botones con 1 demo. "
        "A veces selección accidental con el lápiz. Le gustaría tener matemáticas.",
    ),
    # ── Oliver (11:39-12:08) ──────────────────────────────────────────────
    (
        FECHA, SESION, "Oliver", "Individual", "", "OLIBOT", "3", "Verde",
        "Sí", 29, 0, 5, 5,
        "Línea recta, línea curva, l, 2, 3",
        2, "Técnico",
        "Sí (vuelve espontáneamente a 2 y 3)",
        "Contento", "Sí", "Alta", "No usó",
        "3 años — el que más aguanta (29 min). Le flipan las animaciones y los "
        "círculos verdes. Entiende la goma y el sistema de estrellas. Vuelve a "
        "números 2 y 3 por iniciativa propia. 2 min de pintura al final.",
    ),
]

# ── Hoja "Comentarios" ───────────────────────────────────────────────────────
# Columnas: Fecha | Apodo | Tema | Comentario

COMENTARIOS = [
    # ── Héctor ───────────────────────────────────────────────────────────
    (FECHA, "Héctor", "Se le dio bien",
     "Completó 4 ejercicios en el módulo de pintura sin que se lo pidieran. "
     "Recuerda y usa el botón de repetir de forma autónoma. Comparación espontánea "
     "con papel: 'le gusta más OLIBOT, más divertido'."),
    (FECHA, "Héctor", "Le costó / se atascó",
     "No espera el fin del tutorial; lo salta porque 'ya se lo sabe'. "
     "Muy lento al procesar niveles avanzados (azul/rojo). No espera el final "
     "de las animaciones. No sabe dónde está la goma en el módulo de dibujo."),
    (FECHA, "Héctor", "Voz / habla",
     "Quitó el micrófono; no usó la voz para hablar con OLIBOT. "
     "Sin embargo le gustó MUCHO escuchar a OLIBOT hablar (TTS)."),
    (FECHA, "Héctor", "Actitud / emoción",
     "Activo e impaciente con el tutorial; disfruta activamente la pintura. "
     "Le gusta más OLIBOT que las fichas de papel ('más divertido')."),
    (FECHA, "Héctor", "Idea de mejora",
     "1) No repetir tanto '¡Muy bien! Otra vez'. "
     "2) Reducir la frecuencia de preguntas sobre el estado emocional. "
     "3) Añadir opción de dibujo libre. "
     "4) Revisar la claridad del espacio del tutorial."),
    (FECHA, "Héctor", "Incidencia técnica",
     "Quitó el micrófono durante la sesión. "
     "El espacio del tutorial no se entendió a primer vistazo."),

    # ── Cayetana ──────────────────────────────────────────────────────────
    (FECHA, "Cayetana", "Se le dio bien",
     "Pasó todos los niveles de tirón con nivel automático (sin resistencia). "
     "Espera las transiciones y estados de OLIBOT con paciencia. "
     "Contesta las preguntas de OLIBOT y ríe (sin miedo)."),
    (FECHA, "Cayetana", "Le costó / se atascó",
     "Se cansó en la letra 'l' → necesitó cambio de actividad (primero a números, "
     "luego volvió a 'l'). Selecciona el dibujo que le gusta por preferencia personal, "
     "no por categoría curricular."),
    (FECHA, "Cayetana", "Actitud / emoción",
     "Sin miedo. Tranquila y paciente durante toda la sesión. "
     "No se queja aunque ya conoce el contenido → actitud muy positiva."),

    # ── Ángela ────────────────────────────────────────────────────────────
    (FECHA, "Ángela", "Se le dio bien",
     "Atendió muy bien el tutorial la primera vez. "
     "Aprendió todos los botones con una sola demostración del adulto. "
     "Maneja la tablet con soltura (4 años)."),
    (FECHA, "Ángela", "Le costó / se atascó",
     "Mantener el foco en la escritura: a los 5 min (10:40) pasó a pintar "
     "voluntariamente. Selección accidental de elementos con el lápiz/pincel táctil. "
     "Después de la 1ª vez ya no atiende al tutorial (pq se lo sabe)."),
    (FECHA, "Ángela", "Actitud / emoción",
     "Confiada y exploradora. Le gustaría que OLIBOT tuviera matemáticas "
     "(feedback espontáneo de contenido, muy relevante)."),
    (FECHA, "Ángela", "Idea de mejora",
     "Los niños pequeños van directos a pintar cuando tienen la opción → "
     "gestionar la transición escritura→pintura. "
     "Revisar el comportamiento del lápiz táctil en tablet (selección accidental). "
     "Considerar añadir contenido de matemáticas."),

    # ── Oliver ────────────────────────────────────────────────────────────
    (FECHA, "Oliver", "Se le dio bien",
     "Trazos básicos: línea recta, línea curva, 'l'. Entiende perfectamente la goma. "
     "Comprende el sistema de estrellas para ganar. "
     "Vuelve espontáneamente a los números 2 y 3 para seguir practicando."),
    (FECHA, "Oliver", "Actitud / emoción",
     "Muy motivado y entusiasta. Las animaciones y los círculos verdes le 'flipan'. "
     "Le gusta el avatar, el confeti y 'ganar el 10'. "
     "Es el niño que MÁS aguanta de la sesión (29 min) a pesar de ser el más pequeño (3 años)."),
    (FECHA, "Oliver", "Anécdota / frase",
     "'Es el que aguanta.' Contrasta claramente con los niños mayores que se cansaron "
     "antes. A las 12:00 pasó voluntariamente a pintar pero solo duró 2 min. "
     "Dice verbalmente que le gusta (feedback espontáneo)."),
]


def main():
    print(f"Cargando {XLSX}…")
    wb = load_workbook(XLSX)

    # ── Hoja Registro ────────────────────────────────────────────────────────
    ws = wb["Registro"]
    # Buscar la primera fila vacía (columna A)
    first_empty = ws.max_row + 1
    for row in ws.iter_rows(min_row=2, max_col=1):
        if row[0].value is None:
            first_empty = row[0].row
            break

    print(f"  Registro: añadiendo {len(REGISTRO)} filas desde la fila {first_empty}…")
    for i, row_data in enumerate(REGISTRO):
        r = first_empty + i
        for j, val in enumerate(row_data, start=1):
            cell = ws.cell(row=r, column=j, value=val)
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

    # ── Hoja Comentarios ─────────────────────────────────────────────────────
    ws2 = wb["Comentarios"]
    first_empty2 = ws2.max_row + 1
    for row in ws2.iter_rows(min_row=2, max_col=1):
        if row[0].value is None:
            first_empty2 = row[0].row
            break

    print(f"  Comentarios: añadiendo {len(COMENTARIOS)} filas desde la fila {first_empty2}…")
    for i, row_data in enumerate(COMENTARIOS):
        r = first_empty2 + i
        for j, val in enumerate(row_data, start=1):
            cell = ws2.cell(row=r, column=j, value=val)
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

    wb.save(XLSX)
    print(f"OK → {XLSX}")
    print(f"\nResumen:")
    print(f"  · {len(REGISTRO)} niños añadidos a la hoja 'Registro'")
    print(f"  · {len(COMENTARIOS)} entradas añadidas a la hoja 'Comentarios'")


if __name__ == "__main__":
    main()
