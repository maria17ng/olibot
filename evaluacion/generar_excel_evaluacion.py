"""
Genera la plantilla Excel rápida de evaluación de OLIBOT en el aula.

Crea `hoja_rapida_evaluacion.xlsx` con dos hojas:
  1. "Registro"     — una fila por niño o pareja, columnas con lo esencial a medir.
  2. "Comentarios"  — anotaciones cualitativas (anécdotas, frases, incidencias).

Pensado para rellenar a mano/en tablet DURANTE la sesión (rápido, con desplegables)
y luego volcar al cuaderno extendido en casa.

Uso:  ../.venv/bin/python evaluacion/generar_excel_evaluacion.py
"""
from openpyxl import Workbook
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

OUT = "evaluacion/hoja_rapida_evaluacion.xlsx"

# ── Estilos ────────────────────────────────────────────────────────────────
HEADER_FILL = PatternFill("solid", fgColor="4472C4")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=10)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
THIN = Side(style="thin", color="BBBBBB")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
ZEBRA = PatternFill("solid", fgColor="EEF3FB")


def style_header(ws, ncols):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=1, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER
        cell.border = BORDER
    ws.row_dimensions[1].height = 38
    ws.freeze_panes = "A2"


def add_dropdown(ws, col_letter, options, nrows=300):
    dv = DataValidation(
        type="list",
        formula1='"%s"' % ",".join(options),
        allow_blank=True,
    )
    dv.add("%s2:%s%d" % (col_letter, col_letter, nrows + 1))
    ws.add_data_validation(dv)


def fill_body(ws, ncols, nrows=60):
    for r in range(2, nrows + 2):
        for c in range(1, ncols + 1):
            cell = ws.cell(row=r, column=c)
            cell.border = BORDER
            cell.alignment = LEFT
            if r % 2 == 0:
                cell.fill = ZEBRA
        ws.row_dimensions[r].height = 22


wb = Workbook()

# ════════════════════════════════════════════════════════════════════════════
# HOJA 1 — REGISTRO (una fila por niño o pareja)
# ════════════════════════════════════════════════════════════════════════════
ws = wb.active
ws.title = "Registro"

# (encabezado, ancho de columna, opciones de desplegable o None)
COLS = [
    ("Fecha", 11, None),
    ("Sesión nº", 9, None),
    ("Apodo / Iniciales", 16, None),
    ("Modalidad", 12, ["Individual", "Pareja"]),
    ("Pareja con", 14, None),
    ("Grupo", 12, ["OLIBOT", "Papel"]),
    ("Edad", 7, ["3", "4", "5", "6"]),
    ("Nivel (color)", 12, ["Amarillo", "Verde", "Azul", "Rojo"]),
    ("¿1ª vez?", 9, ["Sí", "No"]),
    ("Min. atención", 11, None),
    ("Nº distracciones", 11, None),
    ("Activ. propuestas", 11, None),
    ("Activ. superadas", 11, None),
    ("Contenidos nuevos", 18, None),
    ("Interv. adulto (nº)", 11, None),
    ("Motivo interv.", 14, ["Técnico", "Pedagógico", "Emocional", "Ninguno"]),
    ('"Otra vez" / reintentos', 12, None),
    ("Satisfacción", 12, ["Contento", "Normal", "Triste"]),
    ("¿Repetiría?", 10, ["Sí", "No"]),
    ("Autonomía", 11, ["Alta", "Media", "Baja"]),
    ("¿Usó voz?", 11, ["Bien", "Regular", "Mal", "No usó"]),
    ("Nota rápida", 30, None),
]

for i, (name, _w, _opts) in enumerate(COLS, start=1):
    ws.cell(row=1, column=i, value=name)
    ws.column_dimensions[get_column_letter(i)].width = _w

fill_body(ws, len(COLS))
style_header(ws, len(COLS))

for i, (_name, _w, opts) in enumerate(COLS, start=1):
    if opts:
        add_dropdown(ws, get_column_letter(i), opts)

# ════════════════════════════════════════════════════════════════════════════
# HOJA 2 — COMENTARIOS (cualitativo)
# ════════════════════════════════════════════════════════════════════════════
ws2 = wb.create_sheet("Comentarios")
COLS2 = [
    ("Fecha", 11, None),
    ("Apodo / Iniciales", 16, None),
    ("Tema", 18, ["Se le dio bien", "Le costó / se atascó", "Voz / habla",
                  "Anécdota / frase", "Actitud / emoción", "Incidencia técnica",
                  "Dinámica de pareja", "Idea de mejora"]),
    ("Comentario", 80, None),
]
for i, (name, _w, _opts) in enumerate(COLS2, start=1):
    ws2.cell(row=1, column=i, value=name)
    ws2.column_dimensions[get_column_letter(i)].width = _w

fill_body(ws2, len(COLS2), nrows=80)
style_header(ws2, len(COLS2))
add_dropdown(ws2, "C", COLS2[2][2], nrows=400)
# Comentario: que ajuste el alto al texto
for r in range(2, 82):
    ws2.cell(row=r, column=4).alignment = LEFT

# ════════════════════════════════════════════════════════════════════════════
# HOJA 3 — LEYENDA (referencia rápida)
# ════════════════════════════════════════════════════════════════════════════
ws3 = wb.create_sheet("Leyenda")
leyenda = [
    ("Campo", "Qué anotar"),
    ("Modalidad", "Individual (1 niño) o Pareja (2 niños, turnos alternos)."),
    ("Grupo", "OLIBOT (tablet) o Papel (grupo control con fichas)."),
    ("Nivel (color)", "Amarillo=3a · Verde=4a · Azul=5a · Rojo=6a (al niño solo se le muestra el color)."),
    ("Min. atención", "Minutos de interacción activa con OLIBOT."),
    ("Nº distracciones", "Veces que el niño desvió la atención."),
    ("Activ. propuestas/superadas", "Para la métrica de rendimiento (OLIBOT vs papel)."),
    ("Contenidos nuevos", "Letras/sílabas/trazos nuevos superados en la sesión."),
    ("Interv. adulto", "Nº de ayudas del adulto + motivo (técnico/pedagógico/emocional)."),
    ('"Otra vez" / reintentos', "Engagement voluntario: pidió repetir o reintentó solo."),
    ("Satisfacción", "Encuesta de caras: Contento 😊 / Normal 😐 / Triste 🙁."),
    ("¿Usó voz?", "Si habló con OLIBOT y cómo funcionó el reconocimiento."),
    ("", ""),
    ("Las 6 métricas (presentación)", ""),
    ("1. Atención sostenida", "Min. atención + distracciones."),
    ("2. Rendimiento comparado", "Activ. superadas OLIBOT vs Papel."),
    ("3. Intervención del adulto", "Nº intervenciones + motivo."),
    ("4. Satisfacción", "Caras + ¿repetiría?"),
    ("5. Velocidad de progresión", "Contenidos nuevos por sesión."),
    ("6. Engagement voluntario", '"Otra vez" / reintentos espontáneos.'),
]
ws3.column_dimensions["A"].width = 30
ws3.column_dimensions["B"].width = 75
for r, (a, b) in enumerate(leyenda, start=1):
    ca = ws3.cell(row=r, column=1, value=a)
    cb = ws3.cell(row=r, column=2, value=b)
    cb.alignment = LEFT
    if r == 1 or a in ("Las 6 métricas (presentación)",):
        ca.font = Font(bold=True, color="FFFFFF")
        cb.font = Font(bold=True, color="FFFFFF")
        ca.fill = HEADER_FILL
        cb.fill = HEADER_FILL
    else:
        ca.font = Font(bold=True)

wb.save(OUT)
print("OK ->", OUT)
